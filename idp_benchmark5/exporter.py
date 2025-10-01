import pandas as pd
from pathlib import Path
from datetime import datetime
from typing import Dict, Any
import shutil
import json
import numpy as np
from utils.logger import default_logger
from utils.utils import json_converter

try:
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


logger = default_logger(__name__)


class NumpyJSONEncoder(json.JSONEncoder):
    """
    Custom JSON encoder for numpy types.
    Converts numpy types to their native Python equivalents.
    """
    def default(self, obj):
        if isinstance(obj, np.integer):
            return int(obj)
        if isinstance(obj, np.floating):
            return float(obj)
        if isinstance(obj, np.ndarray):
            return obj.tolist()
        return super(NumpyJSONEncoder, self).default(obj)

class ResultExporter:
    
    EXCEL_COLUMNS = [
        'PDF filename', 'Processing date', 'File Hit Rate (%)',
        'Item line number', 'Item Part number', 'Item description', 'Item Quantity', 'Item Unit',
        'Item details', 'Special instructions',
        'Candidate number',
        'Line No.', 'Stk ID', 'Name', 'UOM Qty', 'UOM', 'Remarks',
        'List Price', 'Line Total', 'Pur Price', 'Pur Acc ID', 'Pur Acc Name',
        'Margin', 'Pur Curr ID',
        'Candidate score', 'Candidate reasons', 'Candidate order date', 'Candidate source file'
    ]

    def __init__(self, base_output_dir: Path):
        """
        Initializes the exporter with a base directory for all outputs.
        """
        self.base_output_dir = Path(base_output_dir)
        self.base_output_dir.mkdir(parents=True, exist_ok=True)
        logger.info(f"ResultExporter initialized. Output will be saved in subdirectories of: {self.base_output_dir}")

    def _prepare_directories(self, source_file_path: Path) -> Path:
        """
        Creates a dedicated subdirectory for a given source file's results.
        e.g., /output/some_file/
        Returns the path to the created subdirectory.
        """
        subdir_name = source_file_path.stem
        output_subdir = self.base_output_dir / subdir_name
        
        output_subdir.mkdir(parents=True, exist_ok=True)
        
        return output_subdir
    
    def save_all_results(
        self,
        source_file_path: Path,
        parsed_data: Dict[str, Any],
        recommendations: Dict[str, Any],
        hit_rate: float
    ):
        try:
            output_subdir = self._prepare_directories(source_file_path)
            
            shutil.copy(source_file_path, output_subdir)

            parsed_data_path = output_subdir / f"{source_file_path.stem}_parsed.json"
            with open(parsed_data_path, "w", encoding='utf-8') as f:
                # Используем энкодер на случай, если в данных от Gemini есть numpy типы
                json.dump(parsed_data, f, indent=2, ensure_ascii=False, cls=NumpyJSONEncoder)

            recommendations_path = output_subdir / f"{source_file_path.stem}_recommendations.json"
            final_recs_json = {
                "metadata": {
                    "source_file": source_file_path.name,
                    "processing_timestamp": datetime.now().isoformat(),
                    "overall_hit_rate_percent": round(hit_rate, 2),
                },
                "recommendations": recommendations
            }
            with open(recommendations_path, "w", encoding='utf-8') as f:
                json.dump(final_recs_json, f, indent=2, ensure_ascii=False, cls=NumpyJSONEncoder)
            
            if OPENPYXL_AVAILABLE:
                excel_output_path = output_subdir / f"{source_file_path.stem}_recommendations.xlsx"
                self._to_excel(
                    pdf_filename=source_file_path.name,
                    parsed_data=parsed_data,
                    recommendations=recommendations,
                    output_excel_path=excel_output_path,
                    hit_rate=hit_rate
                )
            else:
                logger.warning("Skipping Excel export because 'openpyxl' is not installed.")
            
            logger.info(f"Successfully exported all results for '{source_file_path.name}' to {output_subdir}")

        except Exception as e:
            logger.error(f"Error exporting results for file {source_file_path.name}: {e}", exc_info=True)

    def _to_excel(self, *, pdf_filename: str, parsed_data: Dict[str, Any], recommendations: Dict[str, Any], output_excel_path: Path, hit_rate: float):
        """
        Converts a JSON with recommendations into a an XLSX file and applies styling.
        """
        processing_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        all_rows = []

        line_items_map = {str(item.get('line_item_number')): item for item in parsed_data.get('line_items', [])}

        for item_num, candidates in recommendations.items():
            original_item = line_items_map.get(str(item_num), {})
            
            base_data = {
                'PDF filename': pdf_filename, 'Processing date': processing_date,
                'File Hit Rate (%)': round(hit_rate, 2),
                'Item line number': item_num,
                'Item Part number': original_item.get('part_number', ''),
                'Item description': original_item.get('item_description', ''),
                'Item Quantity': original_item.get('quantity', ''),
                'Item Unit': original_item.get('unit', ''),
                'Item details': original_item.get('item_details', ''),
                'Special instructions': original_item.get('special_instructions', ''),
            }

            if not candidates:
                row_data = base_data.copy()
                row_data.update({'Candidate number': 'N/A', 'Name': 'No recommendations found'})
                all_rows.append(row_data)
                continue
            
            for i, cand in enumerate(candidates, 1):
                matched_item = cand.get('matched_item', {})
                row_data = base_data.copy()
                row_data.update({
                    'Candidate number': i,
                    'Line No.': matched_item.get('line_number'), 'Stk ID': matched_item.get('part_number'),
                    'Name': matched_item.get('description'), 'UOM Qty': matched_item.get('quantity'),
                    'UOM': matched_item.get('unit'), 'Remarks': matched_item.get('doc_remarks'),
                    'List Price': matched_item.get('unit_price'), 'Line Total': matched_item.get('line_total'),
                    'Pur Price': matched_item.get('purchase_price'), 'Pur Acc ID': matched_item.get('supplier_id'),
                    'Pur Acc Name': matched_item.get('supplier_name'), 'Margin': matched_item.get('margin'),
                    'Pur Curr ID': matched_item.get('purchase_currency'), 'Candidate score': cand.get('score'),
                    'Candidate reasons': '\n'.join(cand.get('reasons', [])),
                    'Candidate order date': matched_item.get('order_date'),
                    'Candidate source file': matched_item.get('source_file')
                })
                all_rows.append(row_data)

        if not all_rows:
            logger.warning("No data to create the Excel file.")
            return

        df = pd.DataFrame(all_rows)
        df = df.reindex(columns=self.EXCEL_COLUMNS)
        
        with pd.ExcelWriter(output_excel_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Recommendations')
            
            if not OPENPYXL_AVAILABLE:
                logger.warning("openpyxl is not installed. Skipping Excel cell styling.")
                return

            workbook = writer.book
            worksheet = writer.sheets['Recommendations']
            
            try:
                reasons_col_idx = self.EXCEL_COLUMNS.index('Candidate reasons') + 1
                reasons_col_letter = get_column_letter(reasons_col_idx)

                wrap_alignment = Alignment(wrap_text=True, vertical='top')
                
                worksheet.column_dimensions[reasons_col_letter].width = 60

                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet[f"{reasons_col_letter}{row}"]
                    cell.alignment = wrap_alignment

            except (ValueError, KeyError):
                logger.warning("Could not find 'Candidate reasons' column to apply styling.")

        logger.info(f"Excel report with styled cells saved to: {output_excel_path}")
